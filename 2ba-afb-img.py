import queue, threading, json, openpyxl, time, os, io, shelve, requests, tempfile
from datetime import date
from tkinter import *
from tkinter import ttk, Menu, filedialog, messagebox
import urllib.request
from PIL import ImageTk, Image, ImageChops, ImageOps

class noimage_functions:

    def noimage_load(self):
        # Laden noimage bestand
        noimage_file = shelve.open('noimage')
        noimage_set = set()
        noimage_set = noimage_file['noimage']
        self.noimage_lijst.set(value=noimage_set)
        noimage_file.close()
        return(noimage_set)

    def noimage_save(noimage_set):
        # Opslaan noimage bestand
        noimage_file = shelve.open('noimage')
        noimage_file['noimage'] = noimage_set
        noimage_file.close()

    def noimage_compare(self):
        # Laden noimage bestand
        noimage_file = shelve.open('noimage')
        noimage_set = set()
        noimage_set = noimage_file['noimage']
        noimage_set_old = noimage_file['noimage_old']
        result_file = "noimage_output.txt"
        if noimage_set != noimage_set_old:
            with open(result_file,  "w") as fo:
                fo.write("Overzicht nieuwe noimage bestelnummers.\n")
                for i in noimage_set:
                    if i not in noimage_set_old:
                        fo.write(str(i) + '\n')
            fo.close()
            messagebox.showinfo("Peace out!","Nieuwe noimage entries staan in het noimage_output.txt bestand")
        else:
            messagebox.showinfo("Peace out!","Weet je zeker dat je wilt afsluiten? \n Niet dat je een keuze hebt...")
        noimage_file['noimage_old'] = noimage_set

class threadedstuff:
    # --------------------ProductId-2BA---------------------------
    def ProductId2ba(url):
        productId, fab_productId = '', ''
        try:
            if url != 'onbekend':
                r = oauth.get(url)
                if r.status_code == 200:
                    payload = json.loads(r.text)
                    productId = payload['Id']
                    fab_productId = payload['ProductId']
        except:
            pass
        return(productId, fab_productId)
    #-------------------/ProductId-2BA---------------------------
    #---------------------Oauth----------------------------------
    def Oauth2():
        from oauthlib.oauth2 import LegacyApplicationClient
        from requests_oauthlib import OAuth2Session
        global oauth
        global access_token
        ready = PhotoImage(file="img/sign-check.gif")
        not_ready = PhotoImage(file="img/sign-error.gif")
        # needed user info for the token and connection session
        clientid = mainwindow.client_id.get()
        clientsecret = mainwindow.client_secret.get()
        usrname = mainwindow.username.get()
        pssword = mainwindow.password.get()
        tokenurl = mainwindow.token_url.get()

        # The actual token requests which uses the above data
        try:
            oauth = OAuth2Session(client=LegacyApplicationClient(client_id=clientid))
            token = oauth.fetch_token(token_url=tokenurl,
                username=usrname, password=pssword, client_id=clientid,
                client_secret=clientsecret)
            if token:
                mainwindow.token_status.configure(image= ready)
                mainwindow.token_status.image = ready
                return token['access_token'],  oauth
        except Exception as e:      
            messagebox.showinfo("Sapperdeflap", e)
            mainwindow.token_status.configure(image= not_ready)
            mainwindow.token_status.image = not_ready
    #--------------------/Oauth-----------------------------

    def import_module(worker_list=[], *args):
    # import crawler module if available
        import importlib
        modulenaam = worker_list[4]
        for ch in [' ','//','\\','`','*','_','{','}','[',']','(',')','>','#','+','-','.','!','$',':']:
                if ch in modulenaam:
                    modulenaam = modulenaam.replace(ch,'')
        try:
            func = importlib.import_module('ploadies.{}'.format(modulenaam), package=None)
            Afb_module = getattr(func, modulenaam)
            mainwindow.results_var.set('Module gevonden voor leverancier {}.'.format(worker_list[4]))
            return(Afb_module)

        except:
            Afb_module = ''
            return(Afb_module)
    
    def Attachements_module(ArtDict):
        from openpyxl import Workbook
        output = []
        t = time.time()
        wbkred = openpyxl.load_workbook(mainwindow.kred_lijst_loc.get())
        wskred = wbkred.active
        # preparing a list of suppliers for easy searching
        # this list is used when there is no GTIN found
        SupList = []
        supplier_list = set()
        for i in range(2, wskred.max_row+1, 1):
            SupList.append(wskred.cell(row=i,  column=3).value.lower())
            
        counter = wslist_prod.max_row-1
        current = 1
        #for i in range(2, 20, 1): # tijdelijke manier op slechts een gedeelte van een bestand te verwerken
        for obnr in ArtDict:
            brand = ArtDict[obnr]['merk']
            gtin = ArtDict[obnr]['gtin']
            supplier = ArtDict[obnr]['leverancier']
            supplier_list.add(supplier)
            artnr = ArtDict[obnr]['artlev']
            current_brand, gln,  productId, url = "",  "",  "", ""
            #If there is a GTIN present
            if gtin != "" and gtin != None : 
                url = 'https://api.2ba.nl/1/json/Product/DetailsForProduct?gtin=0{}'.format(gtin)
                productId, fab_productId = threadedstuff.ProductId2ba(url)
            # If there is no GTIN present
            if productId == '': 
                #search for valid GLN in kred lijst.xlsx
                if current_brand is not brand and brand != None:
                    for i in range(2, wskred.max_row+1, 1):
                        if wskred.cell(row=i,  column=3).value.lower() == supplier.lower() and wskred.cell(row=i,  column=4).value.lower() == brand.lower():
                            gln = wskred.cell(row=i,  column=5).value
                            current_brand = brand
                if len(str(gln)) > 12:
                    url = 'https://api.2ba.nl/1/json/Product/DetailsByGLNAndProductcodeA?gln={}&productcode={}&includeFeatures=false'.format(gln,  artnr)
                    productId, fab_productId = threadedstuff.ProductId2ba(url)

            if productId == '': 
                for i in range(2, wskred.max_row+1, 1):
                    if wskred.cell(row=i,  column=3).value.lower() == supplier.lower() and wskred.cell(row=i,  column=4).value.lower() == supplier.lower():
                        gln = wskred.cell(row=i,  column=5).value
                url = 'https://api.2ba.nl/1/json/TradeItem/DetailsByGLNAndTradeItemIdA?gln={}&tradeItemId={}'.format(gln, artnr)
                productId, fab_productId = threadedstuff.ProductId2ba(url)
                productId = fab_productId
            # add all data to the ArtDict dictionary
            ArtDict[obnr]['ProductId'] = productId
            ArtDict[obnr]['payload_url'] = url
            ArtDict[obnr]['Afbeelding gevonden'] = 'Nee'
            mainwindow.results_var.set('Ophalen productIds: {}/{}'.format(current, counter))
            current += 1

        for lev in supplier_list:
            current = 1
            counter = 0
            for obnr in ArtDict:
                if ArtDict[obnr]['leverancier'] == lev:
                    counter += 1
            output = []
            titles = ['Attributes', 'Description', 'LanguageCode','PresentationOrder', 'Title', 'Type', 'URL', 'ProductCode',  'Leverancier',  'Bestelnummer', 'TD', 'IH', 'GA', 'MS', 'PF']
            crawlers = mainwindow.add_crawlers.get()
            Attachments = 'https://api.2ba.nl/1/json/Product/AttachmentsA?id='  #Attachments base link
            filters = '&filterTypes=[%22PPI%22, %22PHI%22]'                     #Attachments filters doctype
            ImgFoundCounter = 0
            for obnr in ArtDict:
                if ArtDict[obnr]['leverancier'] == lev:
                    ProductId = ArtDict[obnr]['ProductId']
                    if ProductId != '':
                            url = '{}{}{}'.format(Attachments,  ProductId, filters)
                            try:
                                r = oauth.get(url)
                                if r.status_code == 200:
                                    ploadies = json.loads(r.text)
                                    
                                    # if no attachments are found the product is added to the notFound list and will be excluded from future searches until after 30 days
                                    # see the notfound function above for more details  
                                    if len(ploadies) != 0:
                                        for entry in ploadies:
                                            entry['Artikelnummer'] = ArtDict[obnr]['artlev']
                                            entry['Leverancier'] = ArtDict[obnr]['leverancier']
                                            entry['Bestelnummer'] = obnr
                                            output.append(list(entry.values()))
                                            ArtDict[obnr]['Afbeelding gevonden'] = 'Ja'
                                            ImgFoundCounter += 1
                            except:
                                pass
                    mainwindow.results_var.set('attachments 2ba {}: {}/{}'.format(lev,current, counter))
                    current += 1   
            
            if crawlers == True:
                # crawler begin
                Afb_module = ''
                worker_list = []
                if (len(ArtDict)-ImgFoundCounter) > 0:
                    mainwindow.results_var.set('Crawler voor {} is nu actief.'.format(lev))
                    for i in ArtDict:
                        #if ArtDict[i]['Afbeelding gevonden'] == 'Ja' or ArtDict[i]['Afbeelding gevonden'] == 'Nee': # temp run all command
                        if ArtDict[i]['Afbeelding gevonden'] == 'Nee' and ArtDict[i]['leverancier'] == lev:
                            max_lengte = len(ArtDict[i])
                            Leverancier = ArtDict[i]['leverancier']
                            temp = [ArtDict[i]['artlev'], i, ArtDict[i]['artlev'],ArtDict[i]['gtin'], Leverancier.lower()]
                            worker_list.append(temp)
                if worker_list:
                    Afb_module = threadedstuff.import_module(worker_list[0])
                
                if Afb_module:
                    return_list = Afb_module(worker_list)
                    for item in return_list:
                        if len(item) > max_lengte:
                            max_lengte = len(item)
                
                    if max_lengte > len(titles):
                            for x in range(len(titles), max_lengte, 1):
                            #response = requests.get(return_list[x])
                            #content_type = response.headers['content-type']
                                titles.append('afbeelding {}'.format(x+1))
                        
                    for i in return_list:
                        if i[5] != 'None':
                            temp = ['', '', '', '', i[4], '', i[5], i[0], i[2], i[1], i[6], i[7], i[8], i[9], i[10]]
                            output.append(temp)
                    
                # crawler eind
            if len(output) > 0:
                wbOutput = Workbook()
                wsOutput = wbOutput.active
                wsOutput.append(titles)
                i = 2
                for row in output:
                    j = 1
                    for cell in row:
                        if cell is not None:
                            wsOutput.cell(row=i,  column=j).value = cell
                        else:
                            wsOutput.cell(row=i,  column=j).value = ''
                        j += 1
                    i += 1
                wbOutput.save("\\\\oosterberg.local\\dfs\\Bestandsbeheer\\Logo\\Leverancier foto's preload\\{}-OUTPUT.xlsx".format(lev))
                if os.path.exists("\\\\oosterberg.local\\dfs\\Bestandsbeheer\\Logo\\Leverancier foto's preload\\{}-OUTPUT.xlsx".format(lev)):
                    for item in OPTIONS:
                        if lev in item[:len(lev)]:
                            mainwindow.brand_list.itemconfig(OPTIONS.index(item), bg='yellow')
                            mainwindow.output_label.config(text='Output bestand aanwezig', background='yellow')


        t = (time.time() - t) / 60 
        mainwindow.results_var.set("Het programma is beïndigd na {} minuten. Zie resultaatbestand(en). \n".format(int(t),  lev))
class GUI:
    def close_window(self):
        noimage_functions.noimage_compare(self)
        self.root.destroy()

    def __init__(self):
        self.root = Tk()
        self.root.protocol("WM_DELETE_WINDOW", self.close_window)
        self.root.grid()
        # Laden settings bestand
        self.client_id = StringVar()
        self.client_secret = StringVar()
        self.username = StringVar()
        self.password = StringVar()
        self.token_url = StringVar()
        self.kred_lijst_loc = StringVar()
        self.img_loc = StringVar()
        self.current_image = StringVar()
        self.ext_lijst_input = StringVar()
        self.extension_lijst = StringVar()
        self.youve_been_warned = BooleanVar()
        self.noimage_lijst = StringVar()
        shelve_file = shelve.open('netniet')
        try:
            self.img_loc.set(shelve_file['saved_loc'])
        except:
            pass
        try:
            self.client_id.set(shelve_file['client_id'])
            self.client_secret.set(shelve_file['client_secret'])
            self.username.set(shelve_file['username'])
            self.password.set(shelve_file['password'])
            self.token_url.set(shelve_file['token_url'])
        except:
            self.client_id.set("")
            self.client_secret.set("")
            self.username.set("")
            self.password.set("")
            self.token_url.set("")
        try:
            self.kred_lijst_loc.set(shelve_file['kred_lijst'])
        except:
            self.kred_lijst_loc.set("voer locatie in")
        try:
            ext_lijst = []
            ext_lijst = shelve_file['ext_lijst']
            self.extension_lijst.set(value=ext_lijst)
        except:
            shelve_file['ext_lijst'] = ""
            pass
        shelve_file.close()

        OPTIONS, OPTIONS2, OPTIONS3 = "", "", ""
        self.workflow = StringVar()
        self.leveranciers = StringVar(value=OPTIONS)
        self.add_crawlers = BooleanVar()
        self.add_noimage = BooleanVar()
        self.rem_whspc = BooleanVar(value=True)
        self.root.title("2BA afbeeldingen tool")
        self.root.minsize(500,360)

        self.tab_control = ttk.Notebook(self.root)
        self.tab_control.grid(column=0, row=0)
        self.mainframe = ttk.Frame(self.tab_control, padding="20 5 10 10", borderwidth=3, relief="sunken")
        self.mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
        self.mainframe.columnconfigure(0, weight=1)
        self.mainframe.rowconfigure(14, weight=2)
        self.image_tab = ttk.Frame(self.tab_control,padding="20 5 10 10", borderwidth=5, relief="sunken")
        self.image_tab.grid(column=0, row=0, sticky=(N, W, E, S))
        self.settings_tab = ttk.Frame(self.tab_control,padding="20 5 10 10", borderwidth=5, relief="sunken")
        self.settings_tab.grid(column=0, row=0, sticky=(N, W, E, S))
        self.tab_control.add(self.mainframe, text="Home")
        self.tab_control.add(self.image_tab, text="Images")
        self.tab_control.add(self.settings_tab, text="Settings")
        #------------------------Status-frame-------------------------------------
        self.results_var = StringVar()
        self.results_lbl = ttk.Label(self.root, textvariable=self.results_var, text="")
        self.results_lbl.grid(column=0, row=1, sticky=(N,W,S,E))
        #-----------------------/Status-frame-------------------------------------
        #------------------------Mainframe-tab------------------------------------
        ttk.Label(self.mainframe, text='Afbeelding download tool 2BA').grid(column=0, row=0, sticky=(N,W,E))
        self.btn_open_file = ttk.Button(self.mainframe, text="Bestand selecteren", command=self.pop_dict)
        self.btn_open_file.grid(column=0, row=2, sticky=(N,W))
        ttk.Button(self.mainframe, text='2BA connectie', command=threadedstuff.Oauth2).grid(column=5, row=2, columnspan=4, sticky=(N,W))
        not_ready = PhotoImage(file="img/sign-error.gif")
        self.token_status = ttk.Label(self.mainframe)
        self.token_status.grid(column=10, row=2, sticky=(N,E))
        self.token_status.configure(image= not_ready)
        self.token_status.image = not_ready
        self.brand_list = Listbox(self.mainframe, height=20, listvariable=self.leveranciers, exportselection=False, selectmode='extended', width=50)
        self.brand_list.grid(column=0, row=3, columnspan=5, rowspan=12, sticky=(W,N))
        self.output_label = ttk.Label(self.mainframe, text='', background='')
        self.output_label.grid(column=0, row=15, sticky=(N,W))
        self.btn_auto = ttk.Radiobutton(self.mainframe, variable=self.workflow, text="Select all", value='auto')
        self.btn_man = ttk.Radiobutton(self.mainframe, variable=self.workflow, text="Manual", value='man')
        self.btn_crawl = ttk.Checkbutton(self.mainframe, text="Include crawler", variable=self.add_crawlers, onvalue=True, offvalue=False)
        self.btn_noimage = ttk.Checkbutton(self.mainframe, text="Include noimage", variable=self.add_noimage, onvalue=True, offvalue=False)
        self.btn_start = ttk.Button(self.mainframe, text="Start", command=self.Start)
        self.btn_auto.grid(column=6, row=3, columnspan=1, sticky=(N,W))
        self.btn_man.grid(column=6, row=4, columnspan=1, sticky=(N,W))
        self.btn_crawl.grid(column=6, row=5, columnspan=1, sticky=(N,W))
        self.btn_noimage.grid(column=6, row=6, columnspan=1, sticky=(N,W))
        self.btn_start.grid(column=6, row=13, columnspan=3, sticky=(N,W,E,S))
        self.btn_auto.bind('<Button-1>', self.select_all)
        self.btn_man.bind('<Button-1>', self.select_none)
        self.brand_list.bind('<Button-1>', self.select_listbox)
        #-----------------------/Mainframe-tab-----------------------------------
        #------------------------Images-tab--------------------------------------
        self.found_images = StringVar(value=OPTIONS2)
        self.img_file = StringVar()
        self.var_brand = StringVar()
        self.var_PrCode = StringVar()
        self.var_Bestelnummer = StringVar()
        self.img_dropdown = StringVar(value=OPTIONS3)
        ttk.Label(self.image_tab, text='Afbeelding controle tool').grid(column=0, row=0, sticky=(N,W,E))
        self.btn_open_file = ttk.Button(self.image_tab, text="Bestand selecteren", command=self.pop_dict_img)
        self.btn_open_file.grid(column=0, row=2, sticky=(N,W))
        self.img_list = Listbox(self.image_tab, listvariable=self.found_images, exportselection=False, selectmode='extended', width=14, height=18)
        self.img_list.grid(column=0, row=3, columnspan=2, rowspan=18, sticky=(W,N))
        self.btn_save_single = ttk.Button(self.image_tab, text="Save image", command=self.save_single_img, width=14)
        self.btn_save_single.grid(column=0, row=21, sticky=(N,W))
        self.btn_save_all = ttk.Button(self.image_tab, text="Save all", command=self.save_all_img, width=14)
        self.btn_save_all.grid(column=0, row=22, sticky=(N,W))
        self.lbl_Attr = ttk.Label(self.image_tab, textvariable=self.var_brand)
        self.lbl_PrCode = ttk.Label(self.image_tab, textvariable=self.var_PrCode)
        self.lbl_Attr.grid(column=3, row=2, sticky=W)
        self.lbl_PrCode.grid(column=3, row=3, sticky=W)
        self.btn_noimage_save = ttk.Button(self.image_tab, text="noimage", command=self.save_noimage, width=12)
        self.btn_noimage_save.grid(column=4, row=3, sticky=E)
        self.cb_img = ttk.Combobox(self.image_tab, values=self.img_dropdown, width=80)
        self.cb_img.grid(column=3, row=4, columnspan=2, sticky=W)
        self.btn_whspc = ttk.Checkbutton(self.image_tab, text="Whitespace", variable=self.rem_whspc, onvalue=True, offvalue=False)
        self.btn_whspc.grid(column=3, row=5, sticky=W)
        self.img_display = PhotoImage(file="")
        self.img_view = ttk.Label(self.image_tab)
        self.img_view.grid(column=3, row=6, rowspan=16, sticky=(N,W))
        self.img_view.configure(image= self.img_display)
        self.img_view.image = self.img_display
        self.img_list.bind('<<ListboxSelect>>', self.select_imgbox)
        self.cb_img.bind('<<ComboboxSelected>>', self.select_url)
        #-----------------------/Images-tab--------------------------------------
        #------------------------Settings-tab------------------------------------
        ttk.Label(self.settings_tab, text="Krediteuren bestand").grid(column=1, row=1,columnspan=4, sticky=(N,W,S,E))
        self.Ekred_lijst = ttk.Entry(self.settings_tab, textvariable=self.kred_lijst_loc, width=70)
        self.Ekred_lijst.grid(column=1, row=2, columnspan=8, sticky=(N,W,S,E))
        ttk.Label(self.settings_tab, text="Extensielijst").grid(column=1, row=3,columnspan=4, sticky=(N,W,S,E))
        self.Ext_lijst_entry = ttk.Entry(self.settings_tab, textvariable=self.ext_lijst_input, width=70)
        self.Ext_lijst_entry.grid(column=1, row=4, columnspan=8, sticky=(N,W,S,E))
        self.Ext_lijst_box = Listbox(self.settings_tab, listvariable=self.extension_lijst, exportselection=False, selectmode='extended', width=50)
        self.Ext_lijst_box.grid(column=1, row=5, columnspan=8, rowspan=4, sticky=(N,W,S,E))
        self.Ext_lijst_entry.bind('<Return>', self.save_ext_data)

        ttk.Label(self.settings_tab, text="API Oauth2 gegevens").grid(column=1, row=9,columnspan=4, sticky=(N,W,S,E))
        self.Eclient_id = ttk.Entry(self.settings_tab, textvariable=self.client_id, width=30)
        self.Eclient_secret = ttk.Entry(self.settings_tab, textvariable=self.client_secret, width=30)
        self.Eusername = ttk.Entry(self.settings_tab, textvariable=self.username, width=30)
        self.Epassword = ttk.Entry(self.settings_tab, textvariable=self.password, width=30)
        self.Etoken_url = ttk.Entry(self.settings_tab, textvariable=self.token_url, width=30)
        self.Eclient_id.grid(column=3, row=10, columnspan=4, sticky=E)
        self.Eclient_secret.grid(column=3, row=11, columnspan=4, sticky=E)
        self.Eusername.grid(column=3, row=12, columnspan=4, sticky=E)
        self.Epassword.grid(column=3, row=13, columnspan=4, sticky=E)
        self.Etoken_url.grid(column=3, row=14, columnspan=4, sticky=E)
        ttk.Label(self.settings_tab, text="client_id ").grid(column=1, row=10, columnspan=2, sticky=W)
        ttk.Label(self.settings_tab, text="client_secret ").grid(column=1, row=11, columnspan=2, sticky=W)
        ttk.Label(self.settings_tab, text="username ").grid(column=1, row=12, columnspan=2, sticky=W)
        ttk.Label(self.settings_tab, text="password ").grid(column=1, row=13, columnspan=2, sticky=W)
        ttk.Label(self.settings_tab, text="token_url ").grid(column=1, row=14, columnspan=2, sticky=W)
        for child in self.settings_tab.winfo_children(): child.grid_configure(padx=2, pady=2)
        for e in [self.Eclient_id,self.Eclient_secret,self.Eusername,self.Epassword,self.Etoken_url]:
            e.bind('<Button-1>', self.wijzigen_API)
            e.bind('<Return>', self.save_api_data)
        self.Ekred_lijst.bind('<Return>', self.save_api_data)
        #-----------------------/Settings-tab------------------------------------
        #------------------------Exit-Program------------------------------------

        #-----------------------/Exit-Program------------------------------------
    def save_ext_data(self, event):
        shelve_file = shelve.open('netniet')
        ext_lijst = shelve_file['ext_lijst']
        nieuwe_waarde = self.ext_lijst_input.get()
        if ext_lijst != "":
            update_lijst = ext_lijst
        else:
            update_lijst = []
        if "," in nieuwe_waarde:
            update_lijst.append(nieuwe_waarde)

            try:
                shelve_file['ext_lijst'] = update_lijst
                ext_lijst = shelve_file['ext_lijst']
                self.extension_lijst.set(value=ext_lijst)
                shelve_file.close()
            except:
                mainwindow.results_var.set("onjuiste waarde ingegeven!")
                pass
        else:
            messagebox.showinfo("Verkeerd scheidingsteken!", "Alleen de komma mag als scheidingsteken gebruikt worden!")
        

    def wijzigen_API(self, event):
        waarschuwing = self.youve_been_warned.get()
        if waarschuwing == False:
            messagebox.showinfo("Danger Will Robinson!", """
Client_id en Client_secret worden uitsluitend door 2ba aangeleverd!
Alleen wijzigen als je (nieuwe) gegevens bij de hand hebt!
Na wijzigen gegevens opnieuw verbinding maken via knop op hoofdscherm.
""", icon='warning')
        self.youve_been_warned.set(True)
    
    def save_api_data(self, event):
        try:
            shelve_file = shelve.open('netniet')
            shelve_file['kred_lijst'] = self.kred_lijst_loc.get()
            shelve_file['client_id'] = self.client_id.get()
            shelve_file['client_secret'] = self.client_secret.get()
            shelve_file['username'] = self.username.get()
            shelve_file['password'] = self.password.get()
            shelve_file['token_url'] = self.token_url.get()
            #shelve_file['saved_loc'] = i
            shelve_file.close()
            mainwindow.results_var.set("Locatie opgeslagen")
        except:
            mainwindow.results_var.set("onjuiste waarde ingegeven!")
            pass
    
    def select_all(self, event):
        self.brand_list.select_set(0, END)

    def select_none(self, event):
        self.brand_list.selection_clear(0, END)
    
    def select_listbox(self, event):
        self.workflow.set(value = "man")

    def open_file(self):
        if os.path.exists(self.current_image.get()):
            os.remove(self.current_image.get())

        self.img_view.configure(image= "")
        self.img_view.image = ""
        mainwindow.results_var.set("")
        try:
            file_loc = self.img_loc.get()
            if file_loc == "":
                file = filedialog.askopenfilename(initialdir=os.getcwd(), title='Please select a file')
            else:
                file = filedialog.askopenfilename(initialdir=file_loc, title='Please select a file')
        except FileNotFoundError as fnf_error:
            messagebox.showinfo(fnf_error)
            exit()
        if file:
            wblist = openpyxl.load_workbook(file)
            wslist = wblist.active
            file_loc = file[: file.rfind("/")+1].replace("/","\\")
            shelve_file = shelve.open('netniet')
            shelve_file['saved_loc'] = file_loc
            self.img_loc.set(shelve_file['saved_loc'])
            shelve_file.close()
            mainwindow.results_var.set("Bestand geopend")
            return(wslist, file)

    def pop_dict(self):
        global OPTIONS, wslist_prod
        wslist_prod, file = self.open_file()
        totaal_lev, noimage_lev = [], []
        leverancierset = set()
        noimage_file = shelve.open('noimage')
        noimage_set = set()
        noimage_set = noimage_file['noimage']
        for i in range(2, wslist_prod.max_row+1,1):
            if wslist_prod.cell(row=i, column=4).value not in noimage_set:
                totaal_lev.append(wslist_prod.cell(row=i, column=1).value.replace('/', '_'))
                leverancierset.add(wslist_prod.cell(row=i, column=1).value.replace('/', '_'))
            else:
                noimage_lev.append(wslist_prod.cell(row=i, column=1).value.replace('/', '_'))
                leverancierset.add(wslist_prod.cell(row=i, column=1).value.replace('/', '_'))
        OPTIONS = sorted([item + " [{}][{}]".format(totaal_lev.count(item), noimage_lev.count(item)) for item in leverancierset])
        self.leveranciers.set(value=OPTIONS)
        for lev in OPTIONS:
            if os.path.exists("\\\\oosterberg.local\\dfs\\Bestandsbeheer\\Logo\\Leverancier foto's preload\\{}-OUTPUT.xlsx".format(lev[:lev.find('[')-1])):
                self.brand_list.itemconfig(OPTIONS.index(lev), bg='yellow')
                mainwindow.output_label.config(text='Output bestand aanwezig', background='yellow')

    def Start(self):
        global ArtDict
        work_list = []
        ArtDict = {}
        items = self.brand_list.curselection()
        noimage_file = shelve.open('noimage')
        noimage_set = set()
        noimage_set = noimage_file['noimage']
        for item in range(0, len(items),1):
            work_list.append(OPTIONS[int(items[item])][:OPTIONS[int(items[item])].find('[')-1])
        for i in range(2, wslist_prod.max_row+1,1):
            temp_merk = wslist_prod.cell(row=i, column=1).value.replace('/','_')
            # Zoeken inclusief noimage afbeeldingen
            if temp_merk in work_list and self.add_noimage.get() == True: # <-- True = vinkje include noimage aan
                obnr = wslist_prod.cell(row=i, column=4).value
                ArtDict[obnr] = { 'artlev': wslist_prod.cell(row=i,  column=2).value,
                                'gtin': wslist_prod.cell(row=i, column=6).value,
                                'leverancier': wslist_prod.cell(row=i, column=1).value.replace('/','_'),
                                'merk': wslist_prod.cell(row=i, column=53).value[20:]}
            # Zoeken exclusief noimage afbeeldingen
            elif temp_merk in work_list and wslist_prod.cell(row=i, column=4).value not in noimage_set:
                obnr = wslist_prod.cell(row=i, column=4).value
                ArtDict[obnr] = { 'artlev': wslist_prod.cell(row=i,  column=2).value,
                                'gtin': wslist_prod.cell(row=i, column=6).value,
                                'leverancier': wslist_prod.cell(row=i, column=1).value.replace('/','_'),
                                'merk': wslist_prod.cell(row=i, column=53).value[20:]}

        self.proc_exec(threadedstuff.Attachements_module, arg=ArtDict)


    def proc_exec(self, task, arg=None):
        """
        Runs designated function with threading
        """
        if arg is None or arg is "":
            tp = threading.Thread(target=task)
        else:
            tp = threading.Thread(target=task, args=(arg,))
        tp.start()

    def pop_dict_img(self):
        global OPTIONS2, wslist_img
        wslist_img, file = self.open_file()
        self.img_file.set(file[:file.rfind('.')-7])
        totaal_img = []
        bestelnummerset = set()
        for i in range(2, wslist_img.max_row+1,1):
            totaal_img.append(wslist_img.cell(row=i, column=10).value.replace('/', '_'))
            bestelnummerset.add(wslist_img.cell(row=i, column=10).value.replace('/', '_'))
        OPTIONS2 = sorted([item + " [{}]".format(totaal_img.count(item)) for item in bestelnummerset])
        self.found_images.set(value=OPTIONS2)

    def select_imgbox(self, event):
        if os.path.exists(self.current_image.get()):
            os.remove(self.current_image.get())
        counter = 1
        work_list = []
        img_ddlist = set()
        selected_img = self.img_list.curselection()
        for item in range(0, len(selected_img),1):
            work_list.append(OPTIONS2[int(selected_img[item])][:OPTIONS2[int(selected_img[item])].find('[')-1])
        for i in range(2, wslist_img.max_row+1,1):
            temp_bestelnummer = wslist_img.cell(row=i, column=10).value
            if temp_bestelnummer in work_list:
                self.var_brand.set(wslist_img.cell(row=i, column=9).value)
                for ch in ['/','\\','*','<','>',':','"','|','?','&']:
                    if ch in wslist_img.cell(row=i, column=8).value:
                        self.var_PrCode.set(wslist_img.cell(row=i, column=8).value.replace(ch,'_'))
                    else:
                        self.var_PrCode.set(wslist_img.cell(row=i, column=8).value)
                self.var_Bestelnummer.set(temp_bestelnummer)
                img_ddlist.add(wslist_img.cell(row=i, column=7).value)
                counter += 1
        OPTIONS3 = sorted([item for item in img_ddlist])
        self.cb_img['values'] = OPTIONS3
        if len(OPTIONS3) != 1:
            self.img_view.configure(image= "")
            self.img_view.image = ""
            mainwindow.results_var.set("")
        else:
            self.cb_img.set(OPTIONS3[0])
            self.select_url(self.cb_img.get())
  
    def select_url(self, event):
        if os.path.exists(self.current_image.get()):
            os.remove(self.current_image.get())
        Ext_Dict = {}
        url = self.cb_img.get()

        img_resize = self.download_img(url)
        if img_resize.mode == "RGBA":
            color = (255, 255, 255)
            img_resize.load()
            background = Image.new('RGB', img_resize.size, color)
            background.paste(img_resize, mask=img_resize.split()[3])  # 3 is het alpha kanaal (transparantie)
            img_resize = background
        elif img_resize.mode == "LA":
            color = (255, 255, 255)
            img_resize.load()
            background = Image.new('RGB', img_resize.size, color)
            background.paste(img_resize, mask=img_resize.split()[1])  # 3 is het alpha kanaal (transparantie)
            img_resize = background
        if self.rem_whspc.get() != False:
            
            def trim(im):
                # Gebruik getpixel((0,0)) om linksbovenin een pixel te pakken
                # Gebruik getpixel((width-1, height-1)) om rechtsonderin een pixel te pakken
                bg = Image.new(im.mode, im.size, im.getpixel((0,0)))
                BorderColor = im.getpixel((0,0))
                diff = ImageChops.difference(im, bg)
                diff = ImageChops.add(diff, diff, 1.0, -20)
                bbox = diff.getbbox()
                if bbox:
                    im = im.crop(bbox)
                    im = ImageOps.expand(im, border=5, fill=BorderColor)
                    return im

            width, height = img_resize.size
            img_resize = trim(img_resize)
        if img_resize.mode != 'RGB':
            img_resize.convert('RGB')
        img_resize.thumbnail((600,600), Image.ANTIALIAS)
        for ch in ['/','\\','*','<','>',':','"','|','?','&']:
            if ch in self.var_PrCode.get():
                self.var_PrCode.set(self.var_PrCode.get().replace(ch,'_'))         
        try:
            img_resize.save("{}\\{}{}".format(self.img_file.get(), self.var_PrCode.get(), ".jpg"))
        except:
            img_resize.save("{}\\{}{}".format(self.img_file.get(), self.var_PrCode.get(), self.cb_img.get()[-4:]))
        self.img_display = ImageTk.PhotoImage(img_resize)
        mainwindow.results_var.set("Afbeelding {}{} : {}x{}px".format(self.var_PrCode.get(), '.jpg', self.img_display.width(), self.img_display.height()))
        self.img_view.configure(image= self.img_display)
        self.img_view.image = self.img_display
        self.current_image.set("{}\\{}{}".format(self.img_file.get(), self.var_PrCode.get(), ".jpg"))   

    def download_img(self, url):
        shelve_file = shelve.open('netniet')
        ext_lijst = shelve_file['ext_lijst']
        for item in ext_lijst:
            if url[-4:] in item[:4]:
                extensie = item[item.find(",")+1:]
                break
            else:
                extensie = url[-4:]
        buffer = tempfile.SpooledTemporaryFile(max_size=1e9)
        r = requests.get(url, stream=True, headers={'Accept-Encoding': None})
        if r.status_code == 200:
            downloaded = 0
            filesize = int(r.headers['content-length'])
            for chunk in r.iter_content(chunk_size=8192):
                downloaded += len(chunk)
                buffer.write(chunk)
            buffer.seek(0)
            if extensie == '.gif':
                img_resize = Image.open(io.BytesIO(buffer.read())).convert('RGB')
            else:
                img_resize = Image.open(io.BytesIO(buffer.read()))
        return(img_resize)

    def save_single_img(self):
        src_image = self.current_image.get()
        dst_image = src_image.replace("Leverancier foto's preload", "XXX PRODUCTFOTO's")
        os.replace(src_image, dst_image)
        selected_img = self.img_list.curselection()
        self.img_list.delete(int(selected_img[0]))
        OPTIONS2.remove(OPTIONS2[int(selected_img[0])])

    def save_all_img(self):
        del_list = []
        shelve_file = shelve.open('netniet')
        ext_lijst = shelve_file['ext_lijst']
        shelve_file.close()
        for item in OPTIONS2:
            if item[-3:] == "[1]":
                for i in range(2, wslist_img.max_row+1,1):
                    temp_bestelnummer = wslist_img.cell(row=i, column=10).value.replace('/','_')
                    if temp_bestelnummer == item[:item.find('[')-1]:
                        self.var_brand.set(wslist_img.cell(row=i, column=9).value)
                        self.var_PrCode.set(wslist_img.cell(row=i, column=8).value)
                        img_dll = wslist_img.cell(row=i, column=7).value
                        img_resize = self.download_img(img_dll)
                        if img_resize.mode == "RGBA":
                            color = (255, 255, 255)
                            img_resize.load()
                            background = Image.new('RGB', img_resize.size, color)
                            background.paste(img_resize, mask=img_resize.split()[3])  # 3 is het alpha kanaal (transparantie)
                            img_resize = background
                        if self.rem_whspc.get() != False:
                            def trim(im):
                                # Gebruik getpixel((0,0)) om linksbovenin een pixel te pakken
                                # Gebruik getpixel((width-1, height-1)) om rechtsonderin een pixel te pakken
                                bg = Image.new(im.mode, im.size, im.getpixel((0,0)))
                                BorderColor = im.getpixel((0,0))
                                diff = ImageChops.difference(im, bg)
                                diff = ImageChops.add(diff, diff, 1.0, -20)
                                bbox = diff.getbbox()
                                if bbox:
                                    im = im.crop(bbox)
                                    im = ImageOps.expand(im, border=5, fill=BorderColor)
                                    return im

                            width, height = img_resize.size
                            img_resize = trim(img_resize)
                        if img_resize.mode != 'RGB':
                            img_resize.convert('RGB')
                        img_resize.thumbnail((600,600), Image.ANTIALIAS)
                        for ch in ['/','\\','*','<','>',':','"','|','?','&']:
                            if ch in self.var_PrCode.get():
                                self.var_PrCode.set(self.var_PrCode.get().replace(ch,'_'))         
                        img_resize.save("{}\\{}{}".format(self.img_file.get(), self.var_PrCode.get(), ".jpg").replace("Leverancier foto's preload", "XXX PRODUCTFOTO's"))
                        idx = self.img_list.get(0,END).index(item)
                        self.img_list.delete(idx)
                        del_list.append(item)
        for art in del_list:
            OPTIONS2.remove(art)

    def save_noimage(self):
        # Aanroepen huidige noimage set,
        # nieuwe bestelnummer toevoegen en opslaan.
        noimage_set = noimage_functions.noimage_load(self)
        noimage_set.add(self.var_Bestelnummer.get())
        noimage_functions.noimage_save(noimage_set)
        # Geselecteerde entry uit lijst verwijderen
        selected_img = self.img_list.curselection()
        self.img_list.delete(int(selected_img[0]))
        OPTIONS2.remove(OPTIONS2[int(selected_img[0])])
        

if __name__ == "__main__":
    mainwindow =GUI()
    mainwindow.root.mainloop()